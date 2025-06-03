#!/usr/bin/env python3

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill
import subprocess
import os
import threading
class TelegramExcelViewer:
    def __init__(self):
        self.root = tk.Tk()
        self.root.geometry("1200x700")
        
        self.workbook = None
        self.worksheet = None
        self.file_path = None
        self.data = []
        self.current_page = 0
        self.page_size = 20
        self.total_rows = 0
        self.all_data = []
        
        self.default_path = os.path.expanduser("~/Downloads/Porno/Descargar/CanalesUnidos")
        
        self.green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        
        self.setup_ui()
        self.setup_bindings()
    
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        
        load_btn = ttk.Button(button_frame, text="📁 Cargar archivo Excel", command=self.load_file)
        load_btn.grid(row=0, column=0, padx=(0, 10))
        
        ready_btn = ttk.Button(button_frame, text="✅ Ver links listos", command=self.view_ready_links)
        ready_btn.grid(row=0, column=1, padx=(0, 10))
        
        exit_btn = ttk.Button(button_frame, text="❌ Salir", command=self.exit_app)
        exit_btn.grid(row=0, column=2, padx=(0, 10))
        
        self.progress = ttk.Progressbar(button_frame, mode='indeterminate')
        self.progress.grid(row=0, column=3, padx=(10, 0), sticky="ew")
        self.progress.grid_remove()
        
        button_frame.columnconfigure(3, weight=1)
        
        columns = ('Link', 'Formato', 'Duration', 'Size', 'File', 'Text')
        self.tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=20)
        
        for col in columns:
            self.tree.heading(col, text=col)
            if col == 'Link':
                self.tree.column(col, width=300)
            else:
                self.tree.column(col, width=120)
        
        v_scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(main_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.tree.grid(row=1, column=0, sticky="nsew")
        v_scrollbar.grid(row=1, column=1, sticky="ns")
        h_scrollbar.grid(row=2, column=0, sticky="ew")
        
        pagination_frame = ttk.Frame(main_frame)
        pagination_frame.grid(row=3, column=0, sticky="ew", pady=(10, 0))
        
        self.prev_btn = ttk.Button(pagination_frame, text="⬅️ Anterior", command=self.prev_page, state="disabled")
        self.prev_btn.grid(row=0, column=0, padx=(0, 10))
        
        self.page_label = ttk.Label(pagination_frame, text="Página 0 de 0")
        self.page_label.grid(row=0, column=1, padx=(0, 10))
        
        self.next_btn = ttk.Button(pagination_frame, text="Siguiente ➡️", command=self.next_page, state="disabled")
        self.next_btn.grid(row=0, column=2, padx=(0, 10))
        
        self.total_label = ttk.Label(pagination_frame, text="Total: 0 registros")
        self.total_label.grid(row=0, column=3, padx=(10, 0))
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        instructions = ttk.Label(main_frame, 
                               text="📖 Instrucciones: Click en link para abrir • Enter para abrir • Cmd+Click para descargar • Cmd+Enter para descargar • Opt+Click/Opt+Enter para reenviar")
        instructions.grid(row=4, column=0, columnspan=2, pady=(10, 0), sticky="w")
    
    def setup_bindings(self):
        self.tree.bind('<Button-1>', self.on_single_click)
        self.tree.bind('<Return>', self.on_enter_key)
        self.tree.bind('<Command-Return>', self.on_command_enter)
        self.tree.bind('<Option-Return>', self.on_option_enter)
        self.tree.bind('<FocusIn>', lambda _: self.tree.focus_set())
        self.tree.focus_set()
    
    def load_file(self):
        os.makedirs(self.default_path, exist_ok=True)
        
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            initialdir=self.default_path,
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if file_path:
            self.show_progress()
            threading.Thread(target=self._load_file_thread, args=(file_path,), daemon=True).start()
    
    def _load_file_thread(self, file_path):
        try:
            self.file_path = file_path
            self.workbook = openpyxl.load_workbook(file_path)
            self.worksheet = self.workbook.active
            
            self.all_data = []
            if self.worksheet is not None:
                for row_num, row in enumerate(self.worksheet.iter_rows(min_row=2, values_only=True), start=2):
                    if row and row[0]:
                        cell = self.worksheet.cell(row=row_num, column=1)
                        is_green = False
                        if cell and cell.fill and cell.fill.start_color:
                            is_green = (getattr(cell.fill.start_color, "rgb", None) == '90EE90')
                        
                        self.all_data.append({
                            'excel_row': row_num,
                            'data': row,
                            'link': row[0],
                            'is_clicked': is_green
                        })
            
            self.total_rows = len(self.all_data)
            self.current_page = 0
            
            self.root.after(0, self._load_complete)
            
        except Exception as e:
            self.root.after(0, lambda: self._load_error(str(e)))
    
    def _load_complete(self):
        self.hide_progress()
        self.load_current_page()
        self.update_pagination()
        messagebox.showinfo("✅ Éxito", f"Archivo cargado correctamente\n{self.total_rows} registros encontrados")
    
    def _load_error(self, error_msg):
        self.hide_progress()
        messagebox.showerror("❌ Error", f"Error al cargar el archivo: {error_msg}")
    
    def show_progress(self):
        self.progress.grid()
        self.progress.start()
    
    def hide_progress(self):
        self.progress.stop()
        self.progress.grid_remove()
    
    def load_current_page(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.data = []
        
        start_idx = self.current_page * self.page_size
        end_idx = min(start_idx + self.page_size, len(self.all_data))
        
        for i in range(start_idx, end_idx):
            data_item = self.all_data[i]
            row_data = data_item['data']
            
            item_id = self.tree.insert('', tk.END, values=row_data)
            
            if data_item['is_clicked']:
                current_values = list(row_data)
                current_values[0] = f"✅ {data_item['link']}"
                self.tree.item(item_id, values=current_values)
            
            self.data.append({
                'tree_item': item_id,
                'excel_row': data_item['excel_row'],
                'link': data_item['link'],
                'is_clicked': data_item['is_clicked'],
                'all_data_index': i
            })
    
    def update_pagination(self):
        total_pages = (len(self.all_data) + self.page_size - 1) // self.page_size if self.all_data else 0
        current_page_display = self.current_page + 1 if self.all_data else 0
        
        self.page_label.config(text=f"Página {current_page_display} de {total_pages}")
        self.total_label.config(text=f"Total: {len(self.all_data)} registros")
        
        self.prev_btn.config(state="normal" if self.current_page > 0 else "disabled")
        self.next_btn.config(state="normal" if self.current_page < total_pages - 1 else "disabled")
    
    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.load_current_page()
            self.update_pagination()
    
    def next_page(self):
        total_pages = (len(self.all_data) + self.page_size - 1) // self.page_size
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self.load_current_page()
            self.update_pagination()
    
    def on_single_click(self, event):
        item = self.tree.identify('item', event.x, event.y)
        column = self.tree.identify('column', event.x, event.y)
        
        if item and column == '#1':
            if event.state & 0x80000:
                self.forward_link(item)
            elif event.state & 0x8:
                self.download_link(item)
            else:
                self.open_link(item)
    
    def on_enter_key(self, event):
        selection = self.tree.selection()
        if selection:
            if event.state & 0x8:
                self.download_link(selection[0])
            else:
                self.open_link(selection[0])
    
    def on_command_enter(self, _):
        selection = self.tree.selection()
        if selection:
            self.download_link(selection[0])

    def on_option_enter(self, _):
        selection = self.tree.selection()
        if selection:
            self.forward_link(selection[0])

    def forward_link(self, item_id):
        clicked_data = self.get_data_by_item(item_id)
        if clicked_data:
            try:
                self.show_progress()
                threading.Thread(
                    target=self._forward_link_thread, 
                    args=(clicked_data,), 
                    daemon=True
                ).start()

            except Exception as e:
                self.hide_progress()
                messagebox.showerror("❌ Error", f"Error al reenviar: {str(e)}")

    def _forward_link_thread(self, clicked_data):
        try:
            success = self.forward_with_tdl(clicked_data['link'])
            self.root.after(0, lambda: self._forward_complete(success, clicked_data))

        except Exception as e:
            self.root.after(0, lambda: self._forward_error(str(e)))

    def _forward_complete(self, success, clicked_data):
        self.hide_progress()

        if success:
            messagebox.showinfo("✅ Éxito", "Video reenviado correctamente al canal")
            self.mark_as_clicked(clicked_data)
        else:
            messagebox.showerror("❌ Error", "No se pudo reenviar el video ni enviar como texto")

    def _forward_error(self, error_msg):
        self.hide_progress()
        messagebox.showerror("❌ Error", f"Error durante el reenvío: {error_msg}")


    def open_link(self, item_id):
        clicked_data = self.get_data_by_item(item_id)
        if clicked_data:
            try:
                self.open_in_telegram(clicked_data['link'])
                self.mark_as_clicked(clicked_data)
            except Exception as e:
                messagebox.showerror("❌ Error", f"Error al abrir el enlace: {str(e)}")
    
    def download_link(self, item_id):
        clicked_data = self.get_data_by_item(item_id)
        if clicked_data:
            try:
                self.download_with_tlg(clicked_data['link'])
                self.mark_as_clicked(clicked_data)
            except Exception as e:
                messagebox.showerror("❌ Error", f"Error al descargar: {str(e)}")
    
    def get_data_by_item(self, item_id):
        for data_item in self.data:
            if data_item['tree_item'] == item_id:
                return data_item
        return None
    
    def open_in_telegram(self, link):
        try:
            if link.startswith('https://t.me/c/'):
                parts = link.replace('https://t.me/c/', '').split('/')
                if len(parts) >= 2:
                    channel = parts[0]
                    post = parts[1]
                    telegram_link = f"tg://privatepost?channel={channel}&post={post}"
                    subprocess.run(['open', telegram_link])
                else:
                    subprocess.run(['open', link])
            elif link.startswith('tg://privatepost'):
                subprocess.run(['open', link])
            else:
                subprocess.run(['open', link])
        except Exception as e:
            raise Exception(f"Error al abrir el enlace: {str(e)}")
    
    def download_with_tlg(self, link):
        try:
            applescript = f'''
            tell application "iTerm"
                if (count of windows) = 0 then
                    create window with default profile
                end if
                
                tell current window
                    create tab with default profile
                    tell current session
                        write text "tlg 1 {link}"
                    end tell
                end tell
            end tell
            '''
            
            subprocess.run(['osascript', '-e', applescript])
        except Exception as e:
            raise Exception(f"Error al ejecutar tlg: {str(e)}")

    def forward_with_tdl(self, link, data_number=1, target_chat="2532518781"):
        """
        Forward video using tdl forward command with fallback to text message
        """
        try:
            # Get storage path using the same pattern from zshrc functions
            storage_path = os.path.expanduser(f"~/.tdl/oktelegram{data_number}")
            
            # First attempt: Direct forward
            success = self._attempt_direct_forward(link, data_number, target_chat, storage_path)
            
            if not success:
                # Fallback: Send link as text message
                print("🔄 Forward failed, attempting to send as text message...")
                success = self._attempt_send_text(link, data_number, target_chat, storage_path)
            
            return success
            
        except Exception as e:
            raise Exception(f"Error en forward_with_tdl: {str(e)}")
    
    def _attempt_send_text(self, link, data_number, target_chat, storage_path):
        """
        Fallback: Send link as simple text message using tdl forward with text content
        """
        try:
            # Try using tdl forward with text content directly
            # Create a simple text message to forward
            forward_cmd = [
                'tdl', 'forward',
                '--storage', f'type=bolt,path={storage_path}',
                '--from', link,
                '--to', target_chat,
                '--mode', 'clone',  # Use clone mode for text fallback
                '--edit', f'"{link}"'  # Edit to just send the link as text
            ]
            
            print(f"📤 Executing text forward command: {' '.join(forward_cmd)}")
            
            # Execute command with timeout
            result = subprocess.run(
                forward_cmd,
                capture_output=True,
                text=True,
                timeout=30
            )
            
            if result.returncode == 0:
                print("✅ Text message sent successfully!")
                return True
            else:
                print(f"⚠️ Text forward failed, trying alternative method...")
                return self._attempt_send_via_echo(link, data_number, target_chat, storage_path)
                
        except Exception as e:
            print(f"❌ Text forward error: {str(e)}")
            return self._attempt_send_via_echo(link, data_number, target_chat, storage_path)

    def _attempt_direct_forward(self, link, data_number, target_chat, storage_path):
        """
        Attempt to forward using tdl forward command
        """
        try:
            # Build tdl forward command based on documentation
            forward_cmd = [
                'tdl', 'forward',
                '--storage', f'type=bolt,path={storage_path}',
                '--from', link,
                '--to', target_chat,
                '--mode', 'direct'  # Prefer official forward API with auto-fallback
            ]

            print(f"🚀 Executing forward command: {' '.join(forward_cmd)}")

            # Execute command with timeout
            result = subprocess.run(
                forward_cmd,
                capture_output=True,
                text=True,
                timeout=60  # 60 seconds timeout
            )

            if result.returncode == 0:
                print("✅ Forward successful!")
                return True
            else:
                print(f"⚠️ Forward failed with code {result.returncode}")
                print(f"Error output: {result.stderr}")
                return False

        except subprocess.TimeoutExpired:
            print("⏰ Forward command timed out")
            return False
        except Exception as e:
            print(f"❌ Forward command error: {str(e)}")
            return False


    def _attempt_send_via_echo(self, link, data_number, target_chat, storage_path):
        """
        Alternative fallback: Send the link using `echo` and `tdl forward`
        """
        try:
            # Create a temporary file to store the link
            temp_file = os.path.join('/tmp', 'temp_link.txt')
            with open(temp_file, 'w') as f:
                f.write(link)

            # Use echo to pipe the content of the file to tdl forward
            forward_cmd = [
                'tdl', 'forward',
                '--storage', f'type=bolt,path={storage_path}',
                '--from', f'file://{temp_file}',  # Read from temp file
                '--to', target_chat,
                '--mode', 'clone',  # Clone as text
            ]

            print(f"📢 Executing fallback command using echo: {' '.join(forward_cmd)}")

            result = subprocess.run(
                forward_cmd,
                capture_output=True,
                text=True,
                timeout=30  # Timeout
            )

            # Clean up the temporary file
            os.remove(temp_file)

            if result.returncode == 0:
                print("✅ Fallback message sent successfully!")
                return True
            else:
                print(f"❌ Fallback failed, return code: {result.returncode}, error: {result.stderr}")
                return False

        except Exception as e:
            print(f"❌ Error during echo-based fallback: {str(e)}")
            return False

    def mark_as_clicked(self, clicked_data):
        try:
            # Set is_clicked flag
            clicked_data['is_clicked'] = True

            # Update excel if worksheet and workbook are available
            if self.worksheet is not None and self.workbook is not None and self.file_path is not None:
                excel_row = clicked_data['excel_row']
                self.worksheet.cell(row=excel_row, column=1).fill = self.green_fill

                # Save the Excel file
                self.workbook.save(self.file_path)

            # Reload current page to refresh the treeview display
            self.load_current_page()

        except Exception as e:
            messagebox.showerror("❌ Error", f"Error al actualizar Excel: {str(e)}")

    def view_ready_links(self):
        ready_window = tk.Toplevel(self.root)
        ready_window.title("✅ Links Listos")
        ready_window.geometry("600x400")

        ready_frame = ttk.Frame(ready_window, padding="10")
        ready_frame.grid(row=0, column=0, sticky="nsew")

        # Treeview for ready links
        ready_tree = ttk.Treeview(ready_frame, columns=('Link', 'Status'), show='headings')
        ready_tree.heading('Link', text='Link')
        ready_tree.heading('Status', text='Status')
        ready_tree.column('Link', width=450)
        ready_tree.column('Status', width=100)

        # Scrollbars for ready links
        ready_v_scrollbar = ttk.Scrollbar(ready_frame, orient=tk.VERTICAL, command=ready_tree.yview)
        ready_tree.configure(yscrollcommand=ready_v_scrollbar.set)
        
        # Grid layout
        ready_tree.grid(row=0, column=0, sticky="nsew")
        ready_v_scrollbar.grid(row=0, column=1, sticky="ns")
        
        # Configure expansion
        ready_frame.columnconfigure(0, weight=1)
        ready_frame.rowconfigure(0, weight=1)
        ready_window.columnconfigure(0, weight=1)
        ready_window.rowconfigure(0, weight=1)
        
        # Load ready links
        ready_count = 0
        for data_item in self.all_data:
            if data_item['is_clicked']:
                ready_tree.insert('', tk.END, values=(data_item['link'], '✅ Visto'))
                ready_count += 1
        
        # Add refresh button
        refresh_btn = ttk.Button(ready_frame, text="🔄 Actualizar", 
                               command=lambda: self.refresh_ready_links(ready_tree))
        refresh_btn.grid(row=1, column=0, pady=(10, 0), sticky="w")
        
        # Add count label
        count_label = ttk.Label(ready_frame, text=f"Total links vistos: {ready_count}")
        count_label.grid(row=2, column=0, pady=(5, 0), sticky="w")
    
    def refresh_ready_links(self, ready_tree):
        # Clear existing items
        for item in ready_tree.get_children():
            ready_tree.delete(item)
        
        # Reload ready links
        ready_count = 0
        for data_item in self.all_data:
            if data_item['is_clicked']:
                ready_tree.insert('', tk.END, values=(data_item['link'], '✅ Visto'))
                ready_count += 1
        
        messagebox.showinfo("🔄 Actualizado", f"Se encontraron {ready_count} links vistos")
    
    def exit_app(self):
        if messagebox.askokcancel("❌ Salir", "¿Estás seguro de que quieres salir?"):
            self.root.quit()
            self.root.destroy()
    
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = TelegramExcelViewer()
    app.run()
