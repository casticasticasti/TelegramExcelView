#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import PatternFill
import subprocess
import os
import threading
from typing import List, Dict, Any, Optional, Callable


class ExcelHandler:
    """Maneja todas las operaciones relacionadas con archivos Excel"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.file_path = None
        self.green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    
    def load_file(self, file_path: str) -> Dict[str, Any]:
        """
        Carga un archivo Excel y extrae los datos
        
        Args:
            file_path: Ruta del archivo Excel
            
        Returns:
            Dict con información del archivo cargado
        """
        try:
            self.file_path = file_path
            self.workbook = openpyxl.load_workbook(file_path)
            self.worksheet = self.workbook.active
            
            all_data = []
            if self.worksheet is not None:
                for row_num, row in enumerate(self.worksheet.iter_rows(min_row=2, values_only=True), start=2):
                    if row and row[0]:
                        cell = self.worksheet.cell(row=row_num, column=1)
                        is_green = self._is_cell_green(cell)
                        
                        all_data.append({
                            'excel_row': row_num,
                            'data': row,
                            'link': row[0],
                            'is_clicked': is_green
                        })
            
            return {
                'success': True,
                'data': all_data,
                'total_rows': len(all_data),
                'message': f'Archivo cargado correctamente. {len(all_data)} registros encontrados'
            }
            
        except Exception as e:
            return {
                'success': False,
                'data': [],
                'total_rows': 0,
                'message': f'Error al cargar el archivo: {str(e)}'
            }
    
    def _is_cell_green(self, cell) -> bool:
        """Verifica si una celda tiene el formato verde (ya procesada)"""
        if cell and cell.fill and cell.fill.start_color:
            return getattr(cell.fill.start_color, "rgb", None) == '90EE90'
        return False
    
    def mark_as_processed(self, excel_row: int) -> bool:
        """
        Marca una fila como procesada (color verde)
        
        Args:
            excel_row: Número de fila en Excel
            
        Returns:
            True si se marcó correctamente, False en caso contrario
        """
        try:
            if self.worksheet is not None and self.workbook is not None and self.file_path is not None:
                self.worksheet.cell(row=excel_row, column=1).fill = self.green_fill
                self.workbook.save(self.file_path)
                return True
            return False
        except Exception as e:
            print(f"Error al marcar como procesado: {str(e)}")
            return False


class TelegramOperations:
    """Maneja todas las operaciones relacionadas con Telegram"""
    
    @staticmethod
    def open_link(link: str) -> bool:
        """
        Abre un enlace en Telegram
        
        Args:
            link: URL del enlace de Telegram
            
        Returns:
            True si se abrió correctamente, False en caso contrario
        """
        try:
            if link.startswith('https://t.me/c/'):
                parts = link.replace('https://t.me/c/', '').split('/')
                if len(parts) >= 2:
                    channel = parts[0]
                    post = parts[1]
                    telegram_link = f"tg://privatepost?channel={channel}&post={post}"
                    subprocess.run(['open', telegram_link])
                else:
                    subprocess.run(['open', link])
            elif link.startswith('tg://privatepost'):
                subprocess.run(['open', link])
            else:
                subprocess.run(['open', link])
            return True
        except Exception as e:
            print(f"Error al abrir el enlace: {str(e)}")
            return False
    
    @staticmethod
    def download_with_tlg(link: str) -> bool:
        """
        Descarga contenido usando el comando tlg
        
        Args:
            link: URL del contenido a descargar
            
        Returns:
            True si se ejecutó correctamente, False en caso contrario
        """
        try:
            applescript = f'''
            tell application "iTerm"
                if (count of windows) = 0 then
                    create window with default profile
                end if
                
                tell current window
                    create tab with default profile
                    tell current session
                        write text "tlg 1 {link}"
                    end tell
                end tell
            end tell
            '''
            
            subprocess.run(['osascript', '-e', applescript])
            return True
        except Exception as e:
            print(f"Error al ejecutar tlg: {str(e)}")
            return False
    
    @staticmethod
    def forward_with_tdl(link: str, data_number: int = 1, target_chat: str = "2532518781") -> bool:
        """
        Reenvía contenido usando tdl con fallback a mensaje de texto
        
        Args:
            link: URL del contenido a reenviar
            data_number: Número de cuenta de Telegram
            target_chat: ID del chat destino
            
        Returns:
            True si se reenviió correctamente, False en caso contrario
        """
        try:
            storage_path = os.path.expanduser(f"~/.tdl/oktelegram{data_number}")
            
            # Intento directo de reenvío
            success = TelegramOperations._attempt_direct_forward(link, data_number, target_chat, storage_path)
            
            if not success:
                print("🔄 Forward failed, attempting to send as text message...")
                success = TelegramOperations._attempt_send_text(link, data_number, target_chat, storage_path)
            
            return success
            
        except Exception as e:
            print(f"Error en forward_with_tdl: {str(e)}")
            return False
    
    @staticmethod
    def _attempt_direct_forward(link: str, data_number: int, target_chat: str, storage_path: str) -> bool:
        """Intenta reenviar directamente usando tdl forward"""
        try:
            forward_cmd = [
                'tdl', 'forward',
                '--storage', f'type=bolt,path={storage_path}',
                '--from', link,
                '--to', target_chat,
                '--mode', 'direct'
            ]

            print(f"🚀 Executing forward command: {' '.join(forward_cmd)}")

            result = subprocess.run(
                forward_cmd,
                capture_output=True,
                text=True,
                timeout=60
            )

            if result.returncode == 0:
                print("✅ Forward successful!")
                return True
            else:
                print(f"⚠️ Forward failed with code {result.returncode}")
                print(f"Error output: {result.stderr}")
                return False

        except subprocess.TimeoutExpired:
            print("⏰ Forward command timed out")
            return False
        except Exception as e:
            print(f"❌ Forward command error: {str(e)}")
            return False
    
    @staticmethod
    def _attempt_send_text(link: str, data_number: int, target_chat: str, storage_path: str) -> bool:
        """Fallback: Envía el enlace como mensaje de texto"""
        try:
            forward_cmd = [
                'tdl', 'forward',
                '--storage', f'type=bolt,path={storage_path}',
                '--from', link,
                '--to', target_chat,
                '--mode', 'clone',
                '--edit', f'"{link}"'
            ]
            
            print(f"📤 Executing text forward command: {' '.join(forward_cmd)}")
            
            result = subprocess.run(
                forward_cmd,
                capture_output=True,
                text=True,
                timeout=30
            )
            
            if result.returncode == 0:
                print("✅ Text message sent successfully!")
                return True
            else:
                print(f"⚠️ Text forward failed, trying alternative method...")
                return TelegramOperations._attempt_send_via_echo(link, data_number, target_chat, storage_path)
                
        except Exception as e:
            print(f"❌ Text forward error: {str(e)}")
            return TelegramOperations._attempt_send_via_echo(link, data_number, target_chat, storage_path)
    
    @staticmethod
    def _attempt_send_via_echo(link: str, data_number: int, target_chat: str, storage_path: str) -> bool:
        """Método alternativo usando archivo temporal"""
        try:
            temp_file = os.path.join('/tmp', 'temp_link.txt')
            with open(temp_file, 'w') as f:
                f.write(link)

            forward_cmd = [
                'tdl', 'forward',
                '--storage', f'type=bolt,path={storage_path}',
                '--from', f'file://{temp_file}',
                '--to', target_chat,
                '--mode', 'clone',
            ]

            print(f"📢 Executing fallback command using echo: {' '.join(forward_cmd)}")

            result = subprocess.run(
                forward_cmd,
                capture_output=True,
                text=True,
                timeout=30
            )

            os.remove(temp_file)

            if result.returncode == 0:
                print("✅ Fallback message sent successfully!")
                return True
            else:
                print(f"❌ Fallback failed, return code: {result.returncode}, error: {result.stderr}")
                return False

        except Exception as e:
            print(f"❌ Error during echo-based fallback: {str(e)}")
            return False


class DataManager:
    """Maneja la paginación y filtrado de datos"""
    
    def __init__(self, page_size: int = 20):
        self.all_data = []
        self.current_page = 0
        self.page_size = page_size
        self.total_rows = 0
    
    def set_data(self, data: List[Dict[str, Any]]):
        """Establece los datos principales"""
        self.all_data = data
        self.total_rows = len(data)
        self.current_page = 0
    
    def get_current_page_data(self) -> List[Dict[str, Any]]:
        """Obtiene los datos de la página actual"""
        start_idx = self.current_page * self.page_size
        end_idx = min(start_idx + self.page_size, len(self.all_data))
        return self.all_data[start_idx:end_idx]
    
    def get_pagination_info(self) -> Dict[str, Any]:
        """Obtiene información de paginación"""
        total_pages = (len(self.all_data) + self.page_size - 1) // self.page_size if self.all_data else 0
        current_page_display = self.current_page + 1 if self.all_data else 0
        
        return {
            'current_page': current_page_display,
            'total_pages': total_pages,
            'total_records': len(self.all_data),
            'can_go_prev': self.current_page > 0,
            'can_go_next': self.current_page < total_pages - 1
        }
    
    def next_page(self) -> bool:
        """Avanza a la siguiente página"""
        total_pages = (len(self.all_data) + self.page_size - 1) // self.page_size
        if self.current_page < total_pages - 1:
            self.current_page += 1
            return True
        return False
    
    def prev_page(self) -> bool:
        """Retrocede a la página anterior"""
        if self.current_page > 0:
            self.current_page -= 1
            return True
        return False
    
    def get_ready_links(self) -> List[Dict[str, Any]]:
        """Obtiene todos los enlaces que han sido procesados"""
        return [item for item in self.all_data if item['is_clicked']]
    
    def update_item_status(self, all_data_index: int, is_clicked: bool):
        """Actualiza el estado de un elemento"""
        if 0 <= all_data_index < len(self.all_data):
            self.all_data[all_data_index]['is_clicked'] = is_clicked


class AsyncOperationManager:
    """Maneja operaciones asíncronas con callbacks"""
    
    @staticmethod
    def run_async_operation(operation_func: Callable, callback_success: Callable, 
                          callback_error: Callable, *args, **kwargs):
        """
        Ejecuta una operación de forma asíncrona
        
        Args:
            operation_func: Función a ejecutar
            callback_success: Callback en caso de éxito
            callback_error: Callback en caso de error
            *args, **kwargs: Argumentos para operation_func
        """
        def worker():
            try:
                result = operation_func(*args, **kwargs)
                callback_success(result)
            except Exception as e:
                callback_error(str(e))
        
        thread = threading.Thread(target=worker, daemon=True)
        thread.start()


class AppConfig:
    """Configuración de la aplicación"""
    
    DEFAULT_PATH = os.path.expanduser("~/Downloads/Porno/Descargar/CanalesUnidos")
    PAGE_SIZE = 20
    TARGET_CHAT = "2532518781"
    DATA_NUMBER = 1
    
    @staticmethod
    def ensure_default_path():
        """Crea el directorio por defecto si no existe"""
        os.makedirs(AppConfig.DEFAULT_PATH, exist_ok=True)

class TelegramExcelFunctions:
    def __init__(self, config):
        self.excel_handler = ExcelHandler()
        self.telegram_operations = TelegramOperations()
        self.data_manager = DataManager(config['page_size'])
        self.config = config
        self.gui_callback = None

    def load_excel_file(self):
        file_path = self.config.get('default_path')
        if not file_path:
            return False, "No se ha especificado una ruta de archivo"
        result = self.excel_handler.load_file(file_path)
        if result['success']:
            self.data_manager.set_data(result['data'])
            if self.gui_callback:
                self.gui_callback.update_pagination()
                self.gui_callback.load_current_page()
            return True, result['message']
        else:
            return False, result['message']

    def open_in_telegram(self, link):
        self.telegram_operations.open_link(link)
        item = next((item for item in self.data_manager.all_data if item['link'] == link), None)
        if item:
            self.mark_as_clicked(self.data_manager.all_data.index(item))

    def download_with_tlg(self, link):
        self.telegram_operations.download_with_tlg(link)
        item = next((item for item in self.data_manager.all_data if item['link'] == link), None)
        if item:
            self.mark_as_clicked(self.data_manager.all_data.index(item))

    def forward_with_tdl(self, link):
        self.telegram_operations.forward_with_tdl(link, self.config['data_number'], self.config['target_chat'])
        item = next((item for item in self.data_manager.all_data if item['link'] == link), None)
        if item:
            self.mark_as_clicked(self.data_manager.all_data.index(item))

    def get_page_data(self, page_number, page_size):
        return self.data_manager.get_current_page_data()

    def get_total_records(self):
        return self.data_manager.total_rows

    def get_ready_links(self):
        return [item['link'] for item in self.data_manager.get_ready_links()]

    def mark_as_clicked(self, excel_row):
        self.excel_handler.mark_as_processed(excel_row)
        if self.gui_callback:
            self.gui_callback.refresh_current_display()

    def set_gui_callback(self, callback):
        self.gui_callback = callback

    def cleanup(self):
        pass
